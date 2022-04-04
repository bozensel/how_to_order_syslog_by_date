
import csv
import sys, datetime, time
import re
from openpyxl import load_workbook
import openpyxl
import time

def XLSExport(Rows, SheetName, FileName):
    from openpyxl import Workbook
    wb = Workbook()

    ws = wb.active
    ws.title = SheetName
    # ws = wb.create_sheet(SheetName)
    for x in Rows:
        ws.append(x)

    wb.save(FileName)

ExcelExport = [["timestamp","source", "message"]]

rows = []
with open("All-Messages-search-result.csv", 'r') as file:
    csvreader = csv.reader(file)
    header = next(csvreader)
    for row in csvreader:
        row[0].split(";")
        rows.append(row[0])

rows.sort()
        
for i in rows:
    i2 = i.split(";")
    ExcelExport.append([i2[0],i2[1],i2[2]])

XLSExport(ExcelExport, "Messages-result3.xlsx", "Messages-result3.xlsx")

print("'Messages-result3.xlsx' file has been created. You may find sorted syslog files. Thanks for using this program.")
